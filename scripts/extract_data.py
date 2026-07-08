# -*- coding: utf-8 -*-
"""실험결과정리 엑셀 파일에서 운전점 데이터를 추출해 tidy CSV로 저장.

시트 구조: 상단(1~19행)은 연료 물성/헤더, 20행 이후가 운전점 데이터.
각 행 = 1개 운전점 (H2 혼합률 × λ × 점화시기).
사용법: python scripts/extract_data.py <엑셀경로> [출력폴더]
"""
import sys
import openpyxl
import pandas as pd

# 열 번호 -> (컬럼명) 매핑 (r18 헤더 기준)
COLS = {
    1: "test_no", 2: "memo1", 3: "memo2",
    4: "h2_vol_set_pct", 5: "h2_vol_real_pct", 6: "h2_energy_pct",
    7: "speed_rpm", 8: "throttle", 10: "spark_atdc", 11: "map_kpa",
    12: "lambda_meter", 13: "lambda_cal",
    25: "fuel_ng_kgph", 26: "fuel_h2_kgph", 28: "air_kgph",
    35: "vol_eff_pct",
    39: "torque_nm", 40: "power_kw", 41: "bmep_bar", 42: "bte_pct",
    46: "t_intake_c", 47: "t_exh_c",
    77: "speed_comb_rpm", 78: "nimep_bar", 79: "gimep_bar",
    81: "cov_imep_pct", 82: "ca05", 83: "ca10", 84: "ca50", 85: "ca90",
    86: "pmax_bar", 87: "mprr_bar_cad", 89: "burn_dur_cad",
    92: "nite_pct", 93: "gite_pct", 94: "pumping_pct",
    95: "thc_gkwh", 96: "ch4_gkwh", 97: "nmhc_gkwh", 98: "co_gkwh",
    99: "nox_gkwh", 100: "co2_gkwh",
    107: "nox_ppm_raw", 114: "nox_ppm_wet",
    118: "comb_eff_pct",
}

# (H2 set %, MAP kPa) 그룹 -> 명목 λ (그룹 평균 lambda_cal 반올림 기준)
def nominal_lambda(lam_mean: float) -> float:
    for nom in (1.0, 1.1, 1.2, 1.4, 1.6, 1.8, 1.9):
        if abs(lam_mean - nom) < 0.06:
            return nom
    return round(lam_mean, 1)


def main(xlsx_path: str, outdir: str = "data"):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    records = []
    for r in range(20, ws.max_row + 1):
        speed = ws.cell(r, 7).value
        if not isinstance(speed, (int, float)):
            continue  # 데이터 행이 아님
        rec = {"row": r}
        for c, name in COLS.items():
            v = ws.cell(r, c).value
            rec[name] = v
        records.append(rec)

    df = pd.DataFrame(records)

    # 노킹(kp) 플래그: memo1에 'kp' 표기가 있으면 노킹 발생 운전점
    df["knock_flag"] = df["memo1"].astype(str).str.contains("kp", case=False, na=False)
    # NOx 관련 메모 플래그
    df["nox_note"] = df["memo2"].astype(str).str.contains("nox", case=False, na=False)

    # 점화시기: aTDC(음수) -> bTDC(양수, 진각)
    df["spark_btdc"] = -pd.to_numeric(df["spark_atdc"], errors="coerce")

    # 그룹(λ 명목값) 부여: H2×MAP 그룹별 평균 lambda_cal 기준
    df["lambda_cal"] = pd.to_numeric(df["lambda_cal"], errors="coerce")
    grp = df.groupby(["h2_vol_set_pct", "map_kpa"])["lambda_cal"].transform("mean")
    df["lambda_nom"] = grp.apply(nominal_lambda)

    # 숫자형 변환
    for c in df.columns:
        if c not in ("memo1", "memo2", "knock_flag", "nox_note"):
            df[c] = pd.to_numeric(df[c], errors="coerce")

    df = df.sort_values(["h2_vol_set_pct", "lambda_nom", "spark_btdc"]).reset_index(drop=True)
    df.to_csv(f"{outdir}/experiment_tidy.csv", index=False, encoding="utf-8-sig")

    # 최적점: 각 (H2, λ)에서 노킹 미발생(kp 미표기) 운전점 중 BTE 최대
    ok = df[~df["knock_flag"] & df["bte_pct"].notna()]
    idx = ok.groupby(["h2_vol_set_pct", "lambda_nom"])["bte_pct"].idxmax()
    opt = df.loc[idx].sort_values(["h2_vol_set_pct", "lambda_nom"]).reset_index(drop=True)
    opt.to_csv(f"{outdir}/optimal_points.csv", index=False, encoding="utf-8-sig")

    print(f"운전점 {len(df)}개 추출 (노킹 표기 {df['knock_flag'].sum()}개)")
    print(f"최적점 {len(opt)}개:")
    cols = ["h2_vol_set_pct", "lambda_nom", "spark_btdc", "bte_pct",
            "ca50", "cov_imep_pct", "nox_gkwh", "power_kw"]
    print(opt[cols].to_string(index=False))
    return df, opt


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "data/실험결과정리_raw.xlsx"
    outdir = sys.argv[2] if len(sys.argv) > 2 else "data"
    main(xlsx, outdir)
